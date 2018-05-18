from openpyxl import load_workbook

wb = load_workbook('template1.xlsm')    #open workbook
main_sheet = wb[wb.sheetnames[0]]   #wb.sheetnames returns a list of names, wb[] is a dictionary with key is sheet's name
cells = []  #declare a list to store cells' values

for i in range(1, 16): #40 is a tentative number that should be enough to cover the whole request's table
    columns = [] #a list to store all values in a column
    for j in range(10, 13):
        columns.append(main_sheet.cell(column=i, row=j).value)
    cells.append(columns)

# print(cells)

src = []
dest = []

for column in cells:
    # print(column[0])
    if not src:
        if "IP Address / Netmask" == column[0]:
            src.append(column)
    else:
        if "IP Address / Netmask" == column[0]:
            dest.append(column)

print(src)
src = src[0][1:]
dest = dest[0][1:]

print(src)





""""
for network in column
	network = network.split()
	net_if = network[0]
	mask = network[1]
	mask = check_mask(mask)
	network = brand_env_h/n_location-function-netid_mask


brand =
env =
location =
function =
"""




