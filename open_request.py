from openpyxl import load_workbook
from get_src_dst import *

def open_request(request_file):
    wb = load_workbook(request_file)    #open workbook
    main_sheet = wb[wb.sheetnames[0]]   #wb.sheetnames returns a list of names, wb[] is a dictionary with key is sheet's name
    return get_src_dst(main_sheet)










